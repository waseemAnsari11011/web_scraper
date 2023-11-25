import geocoder
import json
import openpyxl
from urllib.parse import urlparse

def load_country_codes():
    with open('CountryCodes.json', 'r') as json_file:
        return json.load(json_file)

def load_country_codes_length():
    with open('Countries_Phone_Number_Length.json', 'r') as json_file:
        return json.load(json_file)

def get_country_code(location_text):
    location = geocoder.osm(location_text)
    if location and location.json:
        country_code = location.json.get('raw', {}).get('address', {}).get('country_code')
        return country_code
    return None

def get_dialing_code(country_code, country_codes_data):
    for country_info in country_codes_data:
        if country_info.get('code') == country_code.upper():
            return country_info.get('dial_code')
    return None

def extract_country_code_and_name(location_text):
    country_codes_data = load_country_codes()
    country_code = get_country_code(location_text)
    if country_code:
        dialing_code = get_dialing_code(country_code, country_codes_data)
        return country_code, dialing_code
    return None, None

def update_phone_with_dialing_code(workbook_path, sheet_name):
    # country_codes_data = load_country_codes()
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb[sheet_name]

    new_col_idx = sheet.max_column + 1
    new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
    new_col_header = 'Updated Phone'

    sheet.insert_cols(new_col_idx)
    sheet[new_col_letter + '1'] = new_col_header

    new_rows = []  # List to store new rows with unique dialing codes

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        headquarters = row[5]  # Assuming headquarters is in the 6th column
        profile_url = row[10]  # Assuming profile URL is in the 11th column
        dialing_codes = {}  # Dictionary to store dialing codes and their counts

        if headquarters:
            country_code, dialing_code = extract_country_code_and_name(headquarters)
            if dialing_code:
                dialing_codes[dialing_code] = dialing_codes.get(dialing_code, 0) + 1

            parts = [part.strip() for part in headquarters.split(',')]
            for part in parts:
                country_code, dialing_code = extract_country_code_and_name(part)
                if dialing_code:
                    dialing_codes[dialing_code] = dialing_codes.get(dialing_code, 0) + 1

        largest_count = max(dialing_codes.values()) if dialing_codes else 0
        most_common_dialing_codes = [code for code, count in dialing_codes.items() if count == largest_count]

        if not most_common_dialing_codes and profile_url:
            parsed_url = urlparse(profile_url)
            domain_parts = parsed_url.netloc.split('.')
            if domain_parts[0].lower() == 'www':
                most_common_dialing_codes.append('+1')  # Default dialing code for www
            else:
                country_code, dialing_code = extract_country_code_and_name(domain_parts[0].lower())
                if country_code:
                    most_common_dialing_codes.append(dialing_code)

        phone_number = str(row[2])  # Convert to string to handle both strings and integers
        # Existing code for most_common_dialing_codes loop
        country_codes_data = load_country_codes_length()  # Moved the data loading outside the loop
        for dialing_code in most_common_dialing_codes:
            if phone_number is not None:
                # 1. Process the phone number to get only number digits
                cleaned_phone = ''.join(filter(lambda char: char.isdigit() or char == '+', str(phone_number)))

                # 2. Use Countries_Phone_Number_Length.js to get the phoneLength for the given dialing code
                phone_lengths = []  # Initialize phone_lengths as an empty list
                for country_info in country_codes_data:
                    if f"+{country_info.get('phone')}" == dialing_code:
                        phone_length_value = country_info.get('phoneLength')
                        if isinstance(phone_length_value, int):
                            phone_lengths = [phone_length_value]  # If it's an int, create a list
                        else:
                            phone_lengths = phone_length_value  # Assume it's a list
                        break

                if phone_lengths is None:
                    continue  # Skip this iteration if phone_lengths is None

                # print('phone_lengths--->', phone_lengths)

                # Iterate over each phone length and generate cleaned phone numbers
                updated_phones = []
                for phone_length in phone_lengths:
                    

                    if phone_number.startswith('+'):
                        updated_phone = phone_number
                    elif phone_number.startswith('00'):
                        phone_removed_0 = cleaned_phone.lstrip('0')
                        updated_phone = f"+{phone_removed_0}"
                    else:
                        if phone_length and len(cleaned_phone) >= phone_length:
                            cleaned_phone_stripped_from_last = cleaned_phone[-phone_length:]
                            updated_phone = f"{dialing_code} {cleaned_phone_stripped_from_last}"
                            
                    updated_phones.append(updated_phone)

                # Append the cleaned phone numbers to new_rows
                for updated_phone in updated_phones:
                    new_row = list(row)
                    new_row[new_col_idx - 1] = updated_phone
                    new_rows.append(new_row)
    # Append the new rows to the sheet
    for new_row in new_rows:
        sheet.append(new_row)

    wb.save(workbook_path)

workbook_path = 'linkedin_data_c.xlsx'
sheet_name = 'Sheet'  # Update with your sheet name
update_phone_with_dialing_code(workbook_path, sheet_name)

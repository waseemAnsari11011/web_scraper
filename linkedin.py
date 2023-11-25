import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import random
from faker import Faker
from fake_useragent import UserAgent
from dotenv import load_dotenv
from bs4 import BeautifulSoup
import openpyxl
import datetime
from urllib.parse import urlparse, urlunparse, urljoin
# Load environment variables from .env file
load_dotenv()


def setup_driver():
    # Path to the downloaded ChromeDriver executable
    os.environ['PATH'] += r";C:\chromeDriver"  # Add the correct path separator

    # Create a Chrome WebDriver instance
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(options=options)
    return driver


def login_linkedin(driver, username, password):
    # Open LinkedIn login page
    driver.get("https://linkedin.com/uas/login")

    # Wait for the username field to be visible
    username_field = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.ID, "username"))
    )
    username_field.send_keys(username)

    # Enter password
    password_field = driver.find_element(By.ID, "password")
    password_field.send_keys(password)

    # Submit the form
    password_field.submit()


def random_delay():
    # Generates a random delay between 2 to 5 seconds
    delay = random.uniform(3, 5)
    time.sleep(delay)


def delay():
    time.sleep(120)


def get_random_user_agent():
    fake = Faker()
    return fake.user_agent()


def scroll_to_bottom(driver):
    start = time.time()

    # will be used in the while loop
    initialScroll = 0
    finalScroll = 1000

    while True:
        driver.execute_script(
            f"window.scrollTo({initialScroll}, {finalScroll});")
        # this command scrolls the window starting from
        # the pixel value stored in the initialScroll
        # variable to the pixel value stored at the
        # finalScroll variable
        initialScroll = finalScroll
        finalScroll += 1000

        # we will stop the script for 3 seconds so that
        # the data can load
        time.sleep(3)
        # You can change it as per your needs and internet speed

        end = time.time()

        # We will scroll for 20 seconds.
        # You can change it as per your needs and internet speed
        if round(end - start) > 20:
            break


def extract_data(driver, profile_url):
    src = driver.page_source

    # Now using beautiful soup
    soup = BeautifulSoup(src, 'html.parser')

    # Initialize variables with empty strings
    name = ""
    overview = ""
    phone = ""
    industry = ""
    company_size = ""
    Headquarters = ""
    Founded = ""
    Specialties = ""

    # Extracting the HTML of the complete introduction box
    intro_box = soup.find('div', {'class': 'block mt2'})
    # Extracting the HTML of the complete "ember-view" class
    company_details = soup.find('section', {
                                'class': 'artdeco-card org-page-details-module__card-spacing artdeco-card org-about-module__margin-bottom'})

    # Write the HTML content to a text file
    with open('company_details.html', 'w', encoding='utf-8') as file:
        file.write(str(company_details))

    if intro_box:
        name_element = intro_box.find(
            'h1', {'class': 'org-top-card-summary__title'})

        if name_element:
            name = name_element.get_text(strip=True)
            print(f"Name: {name}")
    else:
        print("Introduction box not found.")

    if company_details:
        # Extracting Overview
        overview_element = company_details.find(
            'h2', class_='text-heading-xlarge')
        if overview_element:
            overview = overview_element.find_next(
                'p', class_='break-words white-space-pre-wrap t-black--light text-body-medium').get_text(strip=True)
            print(f"Overview: {overview}")

        dt_elements = company_details.find_all(
            'dt', class_='mb1 text-heading-medium')
        for element in dt_elements:
            if "Phone" in element.get_text():
                phone = element.find_next(
                    'span', {'class': 'link-without-visited-state'}).get_text(strip=True)
                print(f"Phone: {phone}")
            elif "Industry" in element.get_text():
                industry = element.find_next(
                    'dd', class_='mb4').get_text(strip=True)
                print(f"Industry: {industry}")
            elif "Company size" in element.get_text():
                # Find the corresponding dd element
                dd_element = element.find_next_sibling('dd')

                # Extract the text from the dd element's contents
                company_size = dd_element.contents[0].strip()
                print(f"Company size: {company_size}")
            elif "Headquarters" in element.get_text():
                Headquarters = element.find_next(
                    'dd', class_='mb4').get_text(strip=True)
                print(f"Headquarters: {Headquarters}")
            elif "Founded" in element.get_text():
                Founded = element.find_next(
                    'dd', class_='mb4').get_text(strip=True)
                print(f"Founded: {Founded}")

            elif "Specialties" in element.get_text():
                Specialties = element.find_next(
                    'dd', class_='mb4').get_text(strip=True)
                print(f"Specialties: {Specialties}")
                break

    # Store extracted data in a dictionary
        data = {
            "Name": name,
            "Overview": overview,
            "Phone": phone,
            "Industry": industry,
            "Company Size": company_size,
            "Headquarters": Headquarters,
            "Founded": Founded,
            "Specialties": Specialties,
            "Date": datetime.datetime.now().strftime("%d/%m/%Y"),
            "Profile URL": profile_url
        }

        return data
    else:
        print("Company details not found.")
        return None


def extract_profile_urls_from_excel(excel_file):
    profile_urls = []

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, max_col=11, values_only=True):
            profile_url = row[10]  # Assuming the URL is in the first column
            profile_urls.append(profile_url)

    except Exception as e:
        print("Error while extracting profile URLs:", e)

    return profile_urls


if __name__ == "__main__":
    # Retrieve credentials from environment variables
    user_email = os.getenv("EMAIL")
    user_pass = os.getenv("PASSWORD")

    # Set up the WebDriver
    driver = setup_driver()
    excel_file_name = "linkedin_data.xlsx"

    try:
        ua = UserAgent()
        random_user_agent = ua.random

        driver.execute_cdp_cmd(
            "Network.setUserAgentOverride",
            {"userAgent": random_user_agent}
        )
        print('random_user_agent---->', random_user_agent)
        
        all_links = [
    'https://ng.linkedin.com/company/mapway-real-estate-limited/about',
    'https://in.linkedin.com/company/blackrock-real-estate-partner/about',
    'https://ca.linkedin.com/company/ig-for-real-estate-company/about'
]


        
        extracted_from_excel = extract_profile_urls_from_excel(excel_file_name)

       # Create sets to store unique links
        unique_links_set = set()
        excel_links_set = set(extracted_from_excel)

        for link in all_links:
            if link not in excel_links_set and link not in unique_links_set:
                unique_links_set.add(link)

        # Convert the set back to a list
        unique_links = list(unique_links_set)

        print("unique_links-->", unique_links)

        if unique_links:
            # Change user agent to a random one
            

            # Log in to LinkedIn
            login_linkedin(driver, user_email, user_pass)

            random_delay()

            # Create an empty list to store data dictionaries
            all_data = []

            for profile_url in unique_links:
                # Introduce a random delay before performing actions
                random_delay()

                # Perform actions on LinkedIn
                driver.get(profile_url)  # Open LinkedIn profile

                # Scroll to the bottom of the profile
                scroll_to_bottom(driver)

                # Extract data
                extracted_data = extract_data(driver, profile_url)

                if extracted_data:
                    all_data.append(extracted_data)

                # Introduce a random delay before going to the next link
                random_delay()

            if not os.path.exists(excel_file_name):
                wb = openpyxl.Workbook()
                ws = wb.active
                headers = list(all_data[0].keys())
                ws.append(headers)
            else:
                wb = openpyxl.load_workbook(excel_file_name)
                ws = wb.active

            for data_dict in all_data:
                ws.append(list(data_dict.values()))

            wb.save(excel_file_name)
            print(f"Data exported to {excel_file_name}")
        else:
            print("No unique links found.")

    finally:
        # Close the browser window and quit the driver
        driver.quit()
